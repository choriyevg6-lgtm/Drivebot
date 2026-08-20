import os
import re
import time
import queue
import logging
import zipfile
import sqlite3
import threading
import urllib.request
from io import BytesIO
from datetime import datetime
from xml.etree import ElementTree

import telebot
from telebot import types, apihelper

from PIL import Image, ImageDraw, ImageFont
from fpdf import FPDF
from openpyxl import load_workbook

# Katta fayllarni yuborishda ulanish uzilib qolmasligi uchun timeout'larni oshiramiz
apihelper.CONNECT_TIMEOUT = 30
apihelper.READ_TIMEOUT = 180

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================
# SOZLAMALAR
# ============================================================
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "8870505838:AAFaT8OMQGT8u-vXX7uOsMGwfdlG4Wiizng")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FONTS_DIR = os.path.join(BASE_DIR, "fonts")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
DB_FILE = os.path.join(BASE_DIR, "bot_data.db")

MAX_PART_BYTES = 15 * 1024 * 1024  # 15 MB - rasmlardan PDF uchun xavfsiz chegara
SAMPLE_TEXT = "Bizning telegram kanal: @Ajoyib_botlarr"

# Bosh admin(lar)ning Telegram user_id'lari. Kerak bo'lsa shu yerga qo'shing.
ADMIN_IDS = {8243491785}

bot = telebot.TeleBot(TELEGRAM_TOKEN, parse_mode="HTML")
BOT_ID = None  # bot ishga tushganda to'ldiriladi

# ============================================================
# KEEP-ALIVE: Render 15 daqiqalik uyqu-muammosini oldini olish
# ============================================================
KEEP_ALIVE_PORT = int(os.environ.get("PORT", 8080))
KEEP_ALIVE_URL = os.environ.get("RENDER_EXTERNAL_URL")


def start_keep_alive_server():
    try:
        from flask import Flask
    except ImportError:
        logger.warning("[keep-alive] 'flask' o'rnatilmagan, veb-server ishga tushmaydi (lokalda bu normal).")
        return

    app = Flask("keep_alive")

    @app.route("/")
    def home():
        return "Bot tirik!"

    def run_flask():
        app.run(host="0.0.0.0", port=KEEP_ALIVE_PORT)

    threading.Thread(target=run_flask, daemon=True).start()
    logger.info(f"[keep-alive] Veb-server {KEEP_ALIVE_PORT}-portda ishga tushdi.")


def start_self_ping():
    if not KEEP_ALIVE_URL:
        logger.info("[keep-alive] RENDER_EXTERNAL_URL topilmadi, self-ping o'chirilgan (lokalda bu normal).")
        return
    try:
        import requests
    except ImportError:
        logger.warning("[keep-alive] 'requests' o'rnatilmagan, self-ping ishlamaydi.")
        return

    def ping_loop():
        while True:
            time.sleep(600)  # har 10 daqiqada
            try:
                requests.get(KEEP_ALIVE_URL, timeout=10)
                logger.info("[keep-alive] Self-ping muvaffaqiyatli.")
            except Exception as e:
                logger.warning(f"[keep-alive] Ping xatosi: {e}")

    threading.Thread(target=ping_loop, daemon=True).start()
    logger.info(f"[keep-alive] Self-ping ishga tushdi -> {KEEP_ALIVE_URL}")

# ============================================================
# BAZA (SQLite) — foydalanuvchilar va majburiy kanallar
# ============================================================
_db_conn = sqlite3.connect(DB_FILE, check_same_thread=False)
_db_conn.row_factory = sqlite3.Row
_db_lock = threading.Lock()


def init_db():
    with _db_lock:
        _db_conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                telegram_username TEXT,
                created_at TEXT,
                is_active INTEGER DEFAULT 1
            )
        """)
        _db_conn.execute("""
            CREATE TABLE IF NOT EXISTS forced_channels (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_identifier TEXT NOT NULL,
                title TEXT,
                added_by INTEGER,
                created_at TEXT
            )
        """)
        _db_conn.commit()


def get_or_create_user(user_id: int, username: str | None):
    with _db_lock:
        cur = _db_conn.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,))
        exists = cur.fetchone()
        if exists:
            _db_conn.execute(
                "UPDATE users SET is_active=1, telegram_username=? WHERE user_id=?",
                (username, user_id),
            )
        else:
            _db_conn.execute(
                "INSERT INTO users (user_id, telegram_username, created_at, is_active) VALUES (?, ?, ?, 1)",
                (user_id, username, datetime.now().isoformat()),
            )
        _db_conn.commit()


def mark_user_active(user_id: int):
    with _db_lock:
        _db_conn.execute("UPDATE users SET is_active=1 WHERE user_id=?", (user_id,))
        _db_conn.commit()


def mark_user_inactive(user_id: int):
    with _db_lock:
        _db_conn.execute("UPDATE users SET is_active=0 WHERE user_id=?", (user_id,))
        _db_conn.commit()


def get_all_user_ids() -> list[int]:
    with _db_lock:
        cur = _db_conn.execute("SELECT user_id FROM users")
        return [row["user_id"] for row in cur.fetchall()]


def get_all_active_user_ids() -> list[int]:
    with _db_lock:
        cur = _db_conn.execute("SELECT user_id FROM users WHERE is_active=1 OR is_active IS NULL")
        return [row["user_id"] for row in cur.fetchall()]


def get_user_stats() -> tuple[int, int]:
    with _db_lock:
        total = _db_conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"]
        active = _db_conn.execute(
            "SELECT COUNT(*) c FROM users WHERE is_active=1 OR is_active IS NULL"
        ).fetchone()["c"]
        return total, active


def list_forced_channels() -> list[sqlite3.Row]:
    with _db_lock:
        cur = _db_conn.execute("SELECT * FROM forced_channels ORDER BY id")
        return cur.fetchall()


def add_forced_channel(identifier: str, title: str, added_by: int):
    with _db_lock:
        _db_conn.execute(
            "INSERT INTO forced_channels (chat_identifier, title, added_by, created_at) VALUES (?, ?, ?, ?)",
            (identifier, title, added_by, datetime.now().isoformat()),
        )
        _db_conn.commit()


def remove_forced_channel(channel_id: int):
    with _db_lock:
        _db_conn.execute("DELETE FROM forced_channels WHERE id=?", (channel_id,))
        _db_conn.commit()


def is_blocked_error(exception) -> bool:
    text = str(exception).lower()
    return any(k in text for k in ("blocked", "chat not found", "deactivated", "forbidden", "kicked"))


def refresh_user_statuses():
    for uid in get_all_user_ids():
        try:
            bot.send_chat_action(uid, "typing")
            mark_user_active(uid)
        except Exception as e:
            if is_blocked_error(e):
                mark_user_inactive(uid)


# ============================================================
# ADMIN / MAJBURIY OBUNA
# ============================================================
def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def normalize_channel_identifier(raw_text: str):
    text = raw_text.strip()
    for prefix in ("https://t.me/", "http://t.me/", "t.me/"):
        if text.lower().startswith(prefix):
            text = text[len(prefix):]
            break
    text = text.strip("/").strip()
    if not text or text.startswith("+") or text.startswith("joinchat"):
        return None
    if text.startswith("@"):
        text = text[1:]
    return "@" + text if text else None


def is_bot_admin_in_chat(identifier: str) -> bool:
    try:
        member = bot.get_chat_member(identifier, BOT_ID)
        return member.status in ("administrator", "creator")
    except Exception:
        return False


def get_missing_subscriptions(user_id: int) -> list[sqlite3.Row]:
    missing = []
    for ch in list_forced_channels():
        try:
            member = bot.get_chat_member(ch["chat_identifier"], user_id)
            if member.status in ("left", "kicked"):
                missing.append(ch)
        except Exception:
            continue
    return missing


def send_subscription_prompt(chat_id: int, missing: list[sqlite3.Row]):
    markup = types.InlineKeyboardMarkup(row_width=1)
    for ch in missing:
        username = ch["chat_identifier"].lstrip("@")
        markup.add(types.InlineKeyboardButton(text=f"➕ {ch['title'] or ch['chat_identifier']}", url=f"https://t.me/{username}"))
    markup.add(types.InlineKeyboardButton(text="✅ Tekshirish", callback_data="check_subs"))
    bot.send_message(
        chat_id,
        "📢 Botdan foydalanish uchun quyidagi kanal(lar)ga a'zo bo'ling, so'ng \"✅ Tekshirish\" tugmasini bosing:",
        reply_markup=markup,
    )

# ============================================================
# SHRIFTLAR (qo'lyozma uslubidagi) — internetdan avtomatik yuklanadi
# ============================================================
FONTS = {
    "Caveat": {
        "file": "Caveat.ttf",
        "url": "https://raw.githubusercontent.com/googlefonts/caveat/main/fonts/ttf/Caveat-Regular.ttf",
        "joined": False,
    },
    "Marck Script": {
        "file": "MarckScript.ttf",
        "url": "https://raw.githubusercontent.com/google/fonts/main/ofl/marckscript/MarckScript-Regular.ttf",
        "joined": True,
    },
    "Kalam": {
        "file": "Kalam.ttf",
        "url": "https://raw.githubusercontent.com/google/fonts/main/ofl/kalam/Kalam-Regular.ttf",
        "joined": False,
    },
    "Pacifico": {
        "file": "Pacifico.ttf",
        "url": "https://raw.githubusercontent.com/google/fonts/main/ofl/pacifico/Pacifico-Regular.ttf",
        "joined": False,
    },
    "Bad Script": {
        "file": "BadScript.ttf",
        "url": "https://raw.githubusercontent.com/google/fonts/main/ofl/badscript/BadScript-Regular.ttf",
        "joined": True,
    },
    "Comforter": {
        "file": "Comforter.ttf",
        "url": "https://raw.githubusercontent.com/google/fonts/main/ofl/comforter/Comforter-Regular.ttf",
        "joined": True,
    },
}

# Oddiy (yozma bo'lmagan), Lotin+Kirill'ni qo'llab-quvvatlaydigan shrift —
# Word/Excel/PPT -> PDF konvertatsiyasida va namunalarda yorliq sifatida ishlatiladi.
PLAIN_FONT_FILE = "PTSans.ttf"
PLAIN_FONT_URL = "https://raw.githubusercontent.com/google/fonts/main/ofl/ptsans/PT_Sans-Web-Regular.ttf"

# Qog'oz turlari
PAPER_TYPES = {
    "daftar": {"label": "📓 Daftar (chiziqli)", "plain_label": "Daftar (chiziqli)", "file": "notebook_page.jpg", "lines": True},
    "list": {"label": "📄 List (toza varaq)", "plain_label": "List (toza varaq)", "file": "plain_page.jpg", "lines": False},
}

PAPER_LABEL_TO_KEY = {info["label"]: key for key, info in PAPER_TYPES.items()}


def ensure_fonts():
    os.makedirs(FONTS_DIR, exist_ok=True)
    for name, info in FONTS.items():
        path = os.path.join(FONTS_DIR, info["file"])
        if os.path.exists(path):
            continue
        try:
            logger.info(f"Yuklanmoqda: {name} ({info['url']})")
            urllib.request.urlretrieve(info["url"], path)
        except Exception as e:
            logger.warning(f"'{name}' shriftini yuklab bo'lmadi: {e}")

    plain_path = os.path.join(FONTS_DIR, PLAIN_FONT_FILE)
    if not os.path.exists(plain_path):
        try:
            urllib.request.urlretrieve(PLAIN_FONT_URL, plain_path)
        except Exception as e:
            logger.warning(f"Oddiy shriftni yuklab bo'lmadi: {e}")


def send_with_retry(func, *args, attempts: int = 3, **kwargs):
    last_error = None
    for i in range(1, attempts + 1):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            last_error = e
            logger.warning(f"Yuborishda xatolik (urinish {i}/{attempts}): {e}")
            time.sleep(3 * i)
    raise last_error


# ============================================================
# FAYLLARNI O'QISH (tashqi og'ir kutubxonalarsiz, faqat stdlib + openpyxl)
# ============================================================
def read_docx_text(path: str) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as z:
        xml_content = z.read("word/document.xml")
    root = ElementTree.fromstring(xml_content)
    paragraphs = []
    for p in root.iter(f"{{{ns['w']}}}p"):
        texts = [node.text or "" for node in p.iter(f"{{{ns['w']}}}t")]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    return "\n".join(paragraphs)


def read_xlsx_text(path: str) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"--- {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(c.strip() for c in cells):
                lines.append("  |  ".join(cells))
        lines.append("")
    return "\n".join(lines)


def read_pptx_text(path: str) -> str:
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    slides = []
    with zipfile.ZipFile(path) as z:
        names = [n for n in z.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", n)]

        def slide_num(n):
            m = re.search(r"slide(\d+)\.xml$", n)
            return int(m.group(1)) if m else 0

        names.sort(key=slide_num)
        for idx, name in enumerate(names, start=1):
            root = ElementTree.fromstring(z.read(name))
            texts = [node.text or "" for node in root.iter(f"{{{ns_a}}}t")]
            body = "\n".join(t for t in texts if t.strip())
            slides.append(f"--- {idx}-slayd ---\n{body}" if body else f"--- {idx}-slayd ---")
    return "\n\n".join(slides)


DOCUMENT_READERS = {
    ".docx": read_docx_text,
    ".xlsx": read_xlsx_text,
    ".pptx": read_pptx_text,
}


# ============================================================
# QOG'OZ FONI (chiziqli daftar yoki toza list) - dasturiy chiziladi
# ============================================================
def ensure_background(paper_kind: str) -> str:
    info = PAPER_TYPES[paper_kind]
    path = os.path.join(TEMPLATES_DIR, info["file"])
    if os.path.exists(path):
        return path

    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    width, height = 1000, 1414
    bg_color = (255, 253, 245) if info["lines"] else (255, 255, 255)
    img = Image.new("RGB", (width, height), bg_color)

    if info["lines"]:
        draw = ImageDraw.Draw(img)
        y = 130
        while y < height - 50:
            draw.line([(70, y), (width - 50, y)], fill=(190, 200, 230), width=1)
            y += 38
        draw.line([(105, 0), (105, height)], fill=(230, 150, 150), width=2)

    img.save(path, quality=90)
    return path


# ============================================================
# MATNNI QO'LYOZMA SAHIFALARGA CHIZISH (faqat handwriting rejimi uchun)
# ============================================================
def render_pages(
    text: str,
    font_path: str,
    paper_kind: str,
    font_size: int = 34,
    text_color: tuple = (25, 25, 112),
    top_label: str | None = None,
) -> list[Image.Image]:
    base_img = Image.open(ensure_background(paper_kind)).convert("RGB")
    x_start, y_start = 115, 130
    right_margin = 50
    bottom_margin = 50
    line_height = 38
    max_width = base_img.width - right_margin

    font = ImageFont.truetype(font_path, font_size)

    pages = []
    img = base_img.copy()
    draw = ImageDraw.Draw(img)

    if top_label:
        label_font = ImageFont.truetype(os.path.join(FONTS_DIR, PLAIN_FONT_FILE), 20)
        draw.text((20, 18), top_label, font=label_font, fill=(130, 130, 130))

    current_x, current_y = x_start, y_start

    for paragraph in text.split("\n"):
        words = paragraph.split(" ") if paragraph else [""]
        for word in words:
            piece = word + " "
            bbox = font.getbbox(piece)
            piece_width = bbox[2] - bbox[0]

            if current_x + piece_width > max_width:
                current_x = x_start
                current_y += line_height

            if current_y + line_height > base_img.height - bottom_margin:
                pages.append(img)
                img = base_img.copy()
                draw = ImageDraw.Draw(img)
                current_x, current_y = x_start, y_start

            draw.text((current_x, current_y), piece, font=font, fill=text_color)
            current_x += piece_width

        current_x = x_start
        current_y += line_height

    pages.append(img)
    return pages


def make_sample_image(font_path: str, paper_kind: str, font_name: str, paper_plain_label: str) -> BytesIO:
    label = f"Shrift: {font_name} — {paper_plain_label}"
    pages = render_pages(SAMPLE_TEXT, font_path, paper_kind, font_size=40, top_label=label)
    buf = BytesIO()
    pages[0].save(buf, format="JPEG", quality=90)
    buf.seek(0)
    return buf


# ============================================================
# QO'LYOZMA SAHIFALARINI HAJMGA QARAB PDF/JPG QILIB TAYYORLASH
# (faqat rasm-asosli chiqishlar uchun: qo'lyozma va Rasmdan-PDF)
# ============================================================
def export_parts(pages: list[Image.Image], force_pdf: bool = False) -> list[tuple[BytesIO, str, str]]:
    if len(pages) == 1 and not force_pdf:
        buf = BytesIO()
        pages[0].save(buf, format="JPEG", quality=90)
        buf.seek(0)
        return [(buf, "jpg", "")]

    sample_buf = BytesIO()
    pages[0].convert("RGB").save(sample_buf, format="JPEG", quality=90)
    per_page_bytes = max(sample_buf.tell(), 1)

    pages_per_part = max(1, int((MAX_PART_BYTES * 0.85) // per_page_bytes))
    chunks = [pages[i:i + pages_per_part] for i in range(0, len(pages), pages_per_part)]

    results = []
    total_parts = len(chunks)
    for idx, chunk in enumerate(chunks, start=1):
        buf = BytesIO()
        rgb_chunk = [p.convert("RGB") for p in chunk]
        rgb_chunk[0].save(buf, format="PDF", save_all=True, append_images=rgb_chunk[1:])
        buf.seek(0)
        label = f" ({idx}/{total_parts}-qism)" if total_parts > 1 else ""
        results.append((buf, "pdf", label))
    return results


# ============================================================
# HAQIQIY MATNLI (VEKTOR) PDF — Word/Excel/PPT konvertatsiyasi uchun.
# Rasmga aylantirmaydi, shuning uchun fayl juda kichik va shrift bir xilda chiroyli chiqadi.
# ============================================================
def generate_text_pdf(paragraphs: list[str], font_size: int = 13) -> bytes:
    plain_font_path = os.path.join(FONTS_DIR, PLAIN_FONT_FILE)
    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.add_font("PTSans", "", plain_font_path)
    pdf.set_font("PTSans", size=font_size)

    line_height = font_size * 0.55
    for para in paragraphs:
        if para.strip():
            pdf.multi_cell(0, line_height, para, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
        else:
            pdf.ln(line_height / 2)

    return bytes(pdf.output())


def sanitize_filename(name: str) -> str:
    name = re.sub(r"[^\w\-\. ]", "_", name).strip()
    return name[:60] if name else "hujjat"

# ============================================================
# ISH NAVBATI (Job Queue) — 512MB RAM / 0.1 CPU serverda bir vaqtning
# o'zida faqat bitta og'ir vazifa (rasm/PDF yasash) ishlaydi, qolganlari
# navbatda kutadi. Bu xotira tugab, bot butunlay o'chib qolishining oldini oladi.
# ============================================================
job_queue: "queue.Queue" = queue.Queue()


def enqueue_job(chat_id: int, func, *args, **kwargs):
    pending = job_queue.qsize()
    if pending > 0:
        try:
            bot.send_message(chat_id, f"⏳ Navbatga qo'yildingiz ({pending} ta oldingizda). Iltimos kuting...")
        except Exception:
            pass
    job_queue.put((func, args, kwargs))


def job_worker():
    while True:
        func, args, kwargs = job_queue.get()
        try:
            func(*args, **kwargs)
        except Exception:
            logger.exception("Navbatdagi vazifada xatolik")
        finally:
            job_queue.task_done()


# ============================================================
# FOYDALANUVCHI HOLATI
# ============================================================
user_data: dict[int, dict] = {}


def get_settings(user_id: int) -> dict:
    return user_data.setdefault(
        user_id,
        {"font": list(FONTS.keys())[0], "paper": "daftar", "mode": None, "pending_images": []},
    )


# ============================================================
# MENYULAR
# ============================================================
MAIN_MENU_ROW1 = ["🖋 Shrift tanlash", "📄 Qog'oz turi"]
MAIN_MENU_ROW2 = ["🖼 Namuna", "📑 PDF qilish"]
MAIN_MENU_ROW3 = ["ℹ️ Yordam"]
BACK_BUTTON = "⬅️ Orqaga"
ADMIN_BUTTON = "🛠 Admin panel"

# Diqqat: bu tugma matni asosiy menyudagi "📑 PDF qilish" bilan
# TO'QNASHMASLIGI shart — shuning uchun boshqacha nomlangan.
PDF_MENU_DOCX = "📝 Hujjatdan PDF (Word/Excel/PPT)"
PDF_MENU_IMAGES = "🖼 Rasmdan PDF"
PDF_FINISH_BUTTON = "✅ Tayyor — PDF qilish"
CANCEL_BUTTON = "❌ Bekor qilish"
NAME_YES = "✅ Ha"
NAME_NO = "❌ Yo'q"

ADMIN_ADD_CHANNEL = "📢 Kanal qo'shish"
ADMIN_LIST_CHANNELS = "📋 Ulangan kanallar"
ADMIN_BROADCAST = "📣 Xabar tarqatish"
ADMIN_STATS = "📊 Statistika"


def main_menu_keyboard(user_id: int) -> types.ReplyKeyboardMarkup:
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(*MAIN_MENU_ROW1)
    markup.row(*MAIN_MENU_ROW2)
    if is_admin(user_id):
        markup.row(MAIN_MENU_ROW3[0], ADMIN_BUTTON)
    else:
        markup.row(*MAIN_MENU_ROW3)
    return markup


def font_menu_keyboard() -> types.ReplyKeyboardMarkup:
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    names = list(FONTS.keys())
    for i in range(0, len(names), 2):
        markup.row(*names[i:i + 2])
    markup.row(BACK_BUTTON)
    return markup


def paper_menu_keyboard() -> types.ReplyKeyboardMarkup:
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for info in PAPER_TYPES.values():
        markup.row(info["label"])
    markup.row(BACK_BUTTON)
    return markup


def pdf_menu_keyboard() -> types.ReplyKeyboardMarkup:
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(PDF_MENU_DOCX)
    markup.row(PDF_MENU_IMAGES)
    markup.row(BACK_BUTTON)
    return markup


def collecting_images_keyboard() -> types.ReplyKeyboardMarkup:
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(PDF_FINISH_BUTTON)
    markup.row(CANCEL_BUTTON)
    return markup


def yes_no_keyboard() -> types.ReplyKeyboardMarkup:
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(NAME_YES, NAME_NO)
    return markup


def admin_menu_keyboard() -> types.ReplyKeyboardMarkup:
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(ADMIN_ADD_CHANNEL, ADMIN_LIST_CHANNELS)
    markup.row(ADMIN_BROADCAST, ADMIN_STATS)
    markup.row(BACK_BUTTON)
    return markup

# ============================================================
# YORDAMCHI: MAJBURIY OBUNANI TEKSHIRISH
# ============================================================
def blocked_by_subscription(message: types.Message) -> bool:
    missing = get_missing_subscriptions(message.from_user.id)
    if missing:
        send_subscription_prompt(message.chat.id, missing)
        return True
    return False


@bot.callback_query_handler(func=lambda c: c.data == "check_subs")
def cb_check_subs(call: types.CallbackQuery):
    missing = get_missing_subscriptions(call.from_user.id)
    if missing:
        bot.answer_callback_query(call.id, "❗ Hali ham a'zo bo'lmagan kanal(lar) bor.", show_alert=True)
    else:
        bot.answer_callback_query(call.id, "✅ Rahmat! Endi botdan foydalanishingiz mumkin.")
        bot.send_message(call.message.chat.id, "✅ Tayyor! Asosiy menyu:", reply_markup=main_menu_keyboard(call.from_user.id))


# ============================================================
# /start
# ============================================================
@bot.message_handler(commands=["start"])
def cmd_start(message: types.Message):
    get_or_create_user(message.from_user.id, message.from_user.username)
    get_settings(message.from_user.id)
    bot.send_message(
        message.chat.id,
        "👋 Salom! Men matningizni qo'lda yozilgandek qilib rasmga aylantiraman, "
        "shuningdek Word/Excel/PPT va rasmlarni PDF ga aylantirib beraman.\n\n"
        "Pastdagi tugmalardan foydalaning.",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


# ============================================================
# SHRIFT / QOG'OZ / NAMUNA / YORDAM
# ============================================================
@bot.message_handler(func=lambda m: m.text == "🖋 Shrift tanlash")
def open_font_menu(message: types.Message):
    bot.send_message(message.chat.id, "Shriftni tanlang:", reply_markup=font_menu_keyboard())


@bot.message_handler(func=lambda m: m.text in FONTS)
def set_font(message: types.Message):
    get_settings(message.from_user.id)["font"] = message.text
    note = " (harflari bir-biriga ulangan)" if FONTS[message.text]["joined"] else ""
    bot.send_message(
        message.chat.id,
        f"✅ Shrift tanlandi: <b>{message.text}</b>{note}",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


@bot.message_handler(func=lambda m: m.text == "📄 Qog'oz turi")
def open_paper_menu(message: types.Message):
    bot.send_message(message.chat.id, "Qog'oz turini tanlang:", reply_markup=paper_menu_keyboard())


@bot.message_handler(func=lambda m: m.text in PAPER_LABEL_TO_KEY)
def set_paper(message: types.Message):
    key = PAPER_LABEL_TO_KEY[message.text]
    get_settings(message.from_user.id)["paper"] = key
    bot.send_message(
        message.chat.id,
        f"✅ Qog'oz turi: <b>{message.text}</b>",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


@bot.message_handler(func=lambda m: m.text == "ℹ️ Yordam")
def help_menu(message: types.Message):
    bot.send_message(
        message.chat.id,
        "Menga matn yoki .docx fayl yuboring — men uni qo'lyozma ko'rinishidagi rasm/PDF ga aylantirib beraman.\n\n"
        "🖋 <b>Shrift tanlash</b> — yozuv uslubi\n"
        "📄 <b>Qog'oz turi</b> — Daftar yoki List\n"
        "🖼 <b>Namuna</b> — barcha shriftlarning ko'rinishi\n"
        "📑 <b>PDF qilish</b> — Word/Excel/PPT yoki rasmlarni PDF ga aylantirish",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


@bot.message_handler(func=lambda m: m.text == "🖼 Namuna")
def send_samples(message: types.Message):
    if blocked_by_subscription(message):
        return
    ensure_fonts()
    bot.send_message(message.chat.id, "🖼 Namunalar tayyorlanmoqda, biroz kuting...")
    for font_name, info in FONTS.items():
        font_path = os.path.join(FONTS_DIR, info["file"])
        if not os.path.exists(font_path):
            continue
        try:
            daftar_label = PAPER_TYPES["daftar"]["label"]
            list_label = PAPER_TYPES["list"]["label"]
            daftar_buf = make_sample_image(font_path, "daftar", font_name, PAPER_TYPES["daftar"]["plain_label"])
            list_buf = make_sample_image(font_path, "list", font_name, PAPER_TYPES["list"]["plain_label"])
            join_note = " 🔗 ulangan harflar" if info["joined"] else ""
            media = [
                types.InputMediaPhoto(daftar_buf, caption=f"🖋 {font_name}{join_note} — {daftar_label}"),
                types.InputMediaPhoto(list_buf, caption=f"🖋 {font_name}{join_note} — {list_label}"),
            ]
            send_with_retry(bot.send_media_group, message.chat.id, media)
        except Exception as e:
            logger.warning(f"Namuna yaratishda xato ({font_name}): {e}")
    bot.send_message(message.chat.id, "Yoqqan shriftni tanlang 👇", reply_markup=main_menu_keyboard(message.from_user.id))


@bot.message_handler(func=lambda m: m.text == BACK_BUTTON)
def go_back(message: types.Message):
    settings = get_settings(message.from_user.id)
    settings["mode"] = None
    settings["pending_images"] = []
    bot.send_message(message.chat.id, "Asosiy menyu:", reply_markup=main_menu_keyboard(message.from_user.id))


# ============================================================
# PDF QILISH — Hujjatdan PDF (Word/Excel/PPT) / Rasmdan PDF
# ============================================================
@bot.message_handler(func=lambda m: m.text == "📑 PDF qilish")
def open_pdf_menu(message: types.Message):
    if blocked_by_subscription(message):
        return
    bot.send_message(message.chat.id, "Qaysi turini xohlaysiz?", reply_markup=pdf_menu_keyboard())


@bot.message_handler(func=lambda m: m.text == PDF_MENU_DOCX)
def choose_docx_to_pdf(message: types.Message):
    settings = get_settings(message.from_user.id)
    settings["mode"] = "awaiting_document_for_pdf"
    bot.send_message(
        message.chat.id,
        "📝 Menga .docx, .xlsx yoki .pptx faylni yuboring — men uni toza, matnli PDF ga aylantirib beraman.",
        reply_markup=pdf_menu_keyboard(),
    )


@bot.message_handler(func=lambda m: m.text == PDF_MENU_IMAGES)
def choose_images_to_pdf(message: types.Message):
    settings = get_settings(message.from_user.id)
    settings["mode"] = "collecting_images"
    settings["pending_images"] = []
    bot.send_message(
        message.chat.id,
        f"🖼 Rasm(lar)ni yuboring. Barchasini jo'natib bo'lgach, \"{PDF_FINISH_BUTTON}\" tugmasini bosing.",
        reply_markup=collecting_images_keyboard(),
    )


@bot.message_handler(func=lambda m: m.text == CANCEL_BUTTON)
def cancel_pdf_flow(message: types.Message):
    settings = get_settings(message.from_user.id)
    settings["mode"] = None
    settings["pending_images"] = []
    bot.send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=main_menu_keyboard(message.from_user.id))


@bot.message_handler(func=lambda m: m.text == PDF_FINISH_BUTTON)
def finish_images_to_pdf(message: types.Message):
    settings = get_settings(message.from_user.id)
    if settings.get("mode") != "collecting_images" or not settings.get("pending_images"):
        bot.send_message(message.chat.id, "❗ Avval kamida bitta rasm yuboring.")
        return
    settings["mode"] = "awaiting_pdf_name_choice"
    bot.send_message(message.chat.id, "Fayl uchun nom qo'yasizmi?", reply_markup=yes_no_keyboard())


@bot.message_handler(func=lambda m: m.text == NAME_YES and get_settings(m.from_user.id).get("mode") == "awaiting_pdf_name_choice")
def ask_pdf_name(message: types.Message):
    get_settings(message.from_user.id)["mode"] = "awaiting_pdf_name_text"
    bot.send_message(message.chat.id, "PDF nomini yozing (kengaytmasiz):")


@bot.message_handler(func=lambda m: m.text == NAME_NO and get_settings(m.from_user.id).get("mode") == "awaiting_pdf_name_choice")
def skip_pdf_name(message: types.Message):
    filename = f"PDF_{datetime.now().strftime('%Y-%m-%d_%H-%M')}"
    finalize_images_to_pdf(message, filename)


@bot.message_handler(func=lambda m: get_settings(m.from_user.id).get("mode") == "awaiting_pdf_name_text" and not m.text.startswith("/"))
def set_pdf_name(message: types.Message):
    filename = sanitize_filename(message.text)
    finalize_images_to_pdf(message, filename)


def finalize_images_to_pdf(message: types.Message, filename: str):
    settings = get_settings(message.from_user.id)
    images_bytes = settings.get("pending_images", [])
    settings["mode"] = None
    settings["pending_images"] = []
    bot.send_message(message.chat.id, "✍️ PDF tayyorlanmoqda...", reply_markup=main_menu_keyboard(message.from_user.id))
    enqueue_job(message.chat.id, do_images_to_pdf, message.chat.id, images_bytes, filename)


def do_images_to_pdf(chat_id: int, images_bytes: list, filename: str):
    try:
        pil_images = [Image.open(BytesIO(b)).convert("RGB") for b in images_bytes]
        parts = export_parts(pil_images, force_pdf=True)
        total = len(parts)
        for i, (buf, ext, label) in enumerate(parts, start=1):
            buf.name = f"{filename}_{i}.pdf" if total > 1 else f"{filename}.pdf"
            buf.seek(0)
            send_with_retry(bot.send_document, chat_id, buf, caption=f"📑 PDF tayyor{label}")
    except Exception as e:
        logger.exception("Rasmdan PDF yasashda xato")
        bot.send_message(chat_id, f"❌ Xatolik: {e}")


def do_document_to_pdf(chat_id: int, file_bytes: bytes, original_name: str):
    ext = os.path.splitext(original_name)[1].lower()
    reader = DOCUMENT_READERS.get(ext)
    if reader is None:
        bot.send_message(chat_id, "❗ Qo'llab-quvvatlanmaydigan fayl turi.")
        return

    tmp_path = os.path.join(BASE_DIR, f"tmp_doc_{chat_id}_{int(time.time())}{ext}")
    try:
        with open(tmp_path, "wb") as f:
            f.write(file_bytes)

        text = reader(tmp_path)
        if not text.strip():
            bot.send_message(chat_id, "❗ Fayl ichida matn topilmadi.")
            return

        paragraphs = text.split("\n")
        pdf_bytes = generate_text_pdf(paragraphs)

        base_name = sanitize_filename(os.path.splitext(original_name)[0])
        buf = BytesIO(pdf_bytes)
        buf.name = f"{base_name}.pdf"
        size_kb = len(pdf_bytes) / 1024
        send_with_retry(
            bot.send_document, chat_id, buf,
            caption=f"📑 PDF tayyor ({size_kb:.0f} KB)",
        )
    except Exception as e:
        logger.exception("Hujjatni PDF qilishda xato")
        bot.send_message(chat_id, f"❌ Xatolik: {e}")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


# ============================================================
# HUJJAT (.docx / .xlsx / .pptx) QABUL QILISH
# ============================================================
@bot.message_handler(content_types=["document"])
def handle_document(message: types.Message):
    if blocked_by_subscription(message):
        return

    doc = message.document
    name = doc.file_name or ""
    ext = os.path.splitext(name)[1].lower()
    settings = get_settings(message.from_user.id)

    if settings.get("mode") == "awaiting_document_for_pdf":
        if ext not in DOCUMENT_READERS:
            bot.reply_to(message, "❗ Faqat .docx, .xlsx yoki .pptx fayl qabul qilinadi.")
            return
        settings["mode"] = None
        file_info = bot.get_file(doc.file_id)
        file_bytes = bot.download_file(file_info.file_path)
        bot.send_message(message.chat.id, "✍️ PDF tayyorlanmoqda...", reply_markup=main_menu_keyboard(message.from_user.id))
        enqueue_job(message.chat.id, do_document_to_pdf, message.chat.id, file_bytes, name)
        return

    if ext != ".docx":
        bot.reply_to(
            message,
            "❗ Bu fayl turi qo'lyozma rejimida qo'llab-quvvatlanmaydi.\n"
            "PDF ga aylantirish uchun \"📑 PDF qilish\" menyusidan foydalaning.",
        )
        return

    # standart qo'lyozma konvertatsiyasi
    file_info = bot.get_file(doc.file_id)
    file_bytes = bot.download_file(file_info.file_path)
    tmp_path = os.path.join(BASE_DIR, f"tmp_{message.from_user.id}.docx")
    with open(tmp_path, "wb") as f:
        f.write(file_bytes)
    try:
        text = read_docx_text(tmp_path)
    finally:
        os.remove(tmp_path)

    if not text.strip():
        bot.reply_to(message, "❗ Fayl ichida matn topilmadi.")
        return

    process_and_send(message, text)


# ============================================================
# RASM QABUL QILISH (Rasmdan PDF rejimida)
# ============================================================
@bot.message_handler(content_types=["photo"])
def handle_photo(message: types.Message):
    settings = get_settings(message.from_user.id)
    if settings.get("mode") != "collecting_images":
        bot.reply_to(
            message,
            "🖼 Rasmni PDF qilish uchun avval \"📑 PDF qilish\" → \"🖼 Rasmdan PDF\" ni tanlang.",
        )
        return
    if blocked_by_subscription(message):
        return

    file_info = bot.get_file(message.photo[-1].file_id)
    file_bytes = bot.download_file(file_info.file_path)
    settings["pending_images"].append(file_bytes)

    bot.send_message(
        message.chat.id,
        f"✅ Qabul qilindi ({len(settings['pending_images'])} ta rasm). "
        f"Yana rasm yuboring yoki \"{PDF_FINISH_BUTTON}\" tugmasini bosing.",
        reply_markup=collecting_images_keyboard(),
    )

# ============================================================
# ADMIN PANEL
# ============================================================
@bot.message_handler(func=lambda m: m.text == ADMIN_BUTTON and is_admin(m.from_user.id))
def open_admin_menu(message: types.Message):
    bot.send_message(message.chat.id, "🛠 Admin panel:", reply_markup=admin_menu_keyboard())


@bot.message_handler(func=lambda m: m.text == ADMIN_ADD_CHANNEL and is_admin(m.from_user.id))
def admin_add_channel_start(message: types.Message):
    get_settings(message.from_user.id)["mode"] = "admin_awaiting_channel"
    bot.send_message(message.chat.id, "Kanal linkini yuboring (masalan: @kanal yoki https://t.me/kanal):")


@bot.message_handler(func=lambda m: is_admin(m.from_user.id) and get_settings(m.from_user.id).get("mode") == "admin_awaiting_channel" and not m.text.startswith("/"))
def admin_add_channel_finish(message: types.Message):
    settings = get_settings(message.from_user.id)
    settings["mode"] = None

    identifier = normalize_channel_identifier(message.text)
    if not identifier:
        bot.send_message(message.chat.id, "❗ Faqat ochiq (@username) kanal/guruh qo'shish mumkin. Qayta urinib ko'ring.", reply_markup=admin_menu_keyboard())
        return

    if not is_bot_admin_in_chat(identifier):
        bot.send_message(
            message.chat.id,
            f"❗ Bot {identifier} da admin emas. Avval botni o'sha kanalga admin qilib qo'ying, so'ng qayta urinib ko'ring.",
            reply_markup=admin_menu_keyboard(),
        )
        return

    try:
        chat = bot.get_chat(identifier)
        title = chat.title or identifier
    except Exception:
        title = identifier

    add_forced_channel(identifier, title, message.from_user.id)
    bot.send_message(message.chat.id, f"✅ Qo'shildi: {title} ({identifier})", reply_markup=admin_menu_keyboard())


@bot.message_handler(func=lambda m: m.text == ADMIN_LIST_CHANNELS and is_admin(m.from_user.id))
def admin_list_channels(message: types.Message):
    channels = list_forced_channels()
    if not channels:
        bot.send_message(message.chat.id, "Hozircha majburiy kanal yo'q.", reply_markup=admin_menu_keyboard())
        return
    for ch in channels:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"rmchannel_ask|{ch['id']}"))
        bot.send_message(message.chat.id, f"{ch['title'] or ch['chat_identifier']} — {ch['chat_identifier']}", reply_markup=markup)


@bot.callback_query_handler(func=lambda c: c.data.startswith("rmchannel_ask|"))
def cb_rmchannel_ask(call: types.CallbackQuery):
    channel_id = call.data.split("|", 1)[1]
    markup = types.InlineKeyboardMarkup()
    markup.row(
        types.InlineKeyboardButton(text="✅ Ha, o'chir", callback_data=f"rmchannel_confirm|{channel_id}"),
        types.InlineKeyboardButton(text="❌ Bekor qilish", callback_data="rmchannel_cancel"),
    )
    bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=markup)
    bot.answer_callback_query(call.id, "Tasdiqlaysizmi?")


@bot.callback_query_handler(func=lambda c: c.data.startswith("rmchannel_confirm|"))
def cb_rmchannel_confirm(call: types.CallbackQuery):
    channel_id = int(call.data.split("|", 1)[1])
    remove_forced_channel(channel_id)
    bot.edit_message_text("🗑 O'chirildi.", call.message.chat.id, call.message.message_id)
    bot.answer_callback_query(call.id, "O'chirildi.")


@bot.callback_query_handler(func=lambda c: c.data == "rmchannel_cancel")
def cb_rmchannel_cancel(call: types.CallbackQuery):
    bot.answer_callback_query(call.id, "Bekor qilindi.")


@bot.message_handler(func=lambda m: m.text == ADMIN_BROADCAST and is_admin(m.from_user.id))
def admin_broadcast_start(message: types.Message):
    get_settings(message.from_user.id)["mode"] = "admin_awaiting_broadcast"
    bot.send_message(message.chat.id, "Barcha foydalanuvchilarga yuboriladigan xabar matnini kiriting:")


@bot.message_handler(func=lambda m: is_admin(m.from_user.id) and get_settings(m.from_user.id).get("mode") == "admin_awaiting_broadcast" and not m.text.startswith("/"))
def admin_broadcast_preview(message: types.Message):
    settings = get_settings(message.from_user.id)
    settings["mode"] = None
    settings["broadcast_text"] = message.text

    markup = types.InlineKeyboardMarkup()
    markup.row(
        types.InlineKeyboardButton(text="✅ Yuborish", callback_data="bcastconfirm"),
        types.InlineKeyboardButton(text="❌ Bekor qilish", callback_data="bcastcancel"),
    )
    bot.send_message(message.chat.id, f"Ushbu xabar yuborilsinmi?\n\n{message.text}", reply_markup=markup)


@bot.callback_query_handler(func=lambda c: c.data == "bcastconfirm")
def cb_broadcast_confirm(call: types.CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    settings = get_settings(call.from_user.id)
    text = settings.get("broadcast_text")
    settings["broadcast_text"] = None
    bot.answer_callback_query(call.id, "Yuborilmoqda...")
    bot.edit_message_text("📣 Xabar navbatga qo'yildi, yuborilmoqda...", call.message.chat.id, call.message.message_id)
    if text:
        enqueue_job(call.message.chat.id, do_broadcast, call.message.chat.id, text)


@bot.callback_query_handler(func=lambda c: c.data == "bcastcancel")
def cb_broadcast_cancel(call: types.CallbackQuery):
    get_settings(call.from_user.id)["broadcast_text"] = None
    bot.answer_callback_query(call.id, "Bekor qilindi.")
    bot.edit_message_text("❌ Bekor qilindi.", call.message.chat.id, call.message.message_id)


def do_broadcast(chat_id: int, text: str):
    user_ids = get_all_active_user_ids()
    success, failed = 0, 0
    for uid in user_ids:
        try:
            bot.send_message(uid, text)
            success += 1
        except Exception as e:
            failed += 1
            if is_blocked_error(e):
                mark_user_inactive(uid)
        time.sleep(0.05)  # Telegram rate-limit'ga urilmaslik uchun
    bot.send_message(chat_id, f"✅ Yakunlandi: {success} ta yetdi, {failed} ta yetmadi.")


@bot.message_handler(func=lambda m: m.text == ADMIN_STATS and is_admin(m.from_user.id))
def admin_stats(message: types.Message):
    wait_msg = bot.send_message(message.chat.id, "⏳ Tekshirilmoqda...")
    enqueue_job(message.chat.id, do_stats, message.chat.id, wait_msg.message_id)


def do_stats(chat_id: int, wait_message_id: int):
    refresh_user_statuses()
    total, active = get_user_stats()
    try:
        bot.delete_message(chat_id, wait_message_id)
    except Exception:
        pass
    bot.send_message(
        chat_id,
        f"📊 <b>Statistika</b>\n\nJami foydalanuvchi: {total}\nFaol (bot bloklanmagan): {active}\nBloklagan: {total - active}",
    )

# ============================================================
# ODDIY MATN — QO'LYOZMA KONVERTATSIYASI
# ============================================================
@bot.message_handler(content_types=["text"])
def handle_text(message: types.Message):
    if message.text.startswith("/"):
        return
    settings = get_settings(message.from_user.id)
    if settings.get("mode") == "collecting_images":
        bot.send_message(message.chat.id, f"🖼 Iltimos, rasm yuboring yoki \"{PDF_FINISH_BUTTON}\"/\"{CANCEL_BUTTON}\" tugmasini bosing.")
        return
    if settings.get("mode") == "awaiting_document_for_pdf":
        bot.send_message(message.chat.id, "📝 Iltimos, .docx, .xlsx yoki .pptx faylni yuboring.")
        return
    if blocked_by_subscription(message):
        return
    process_and_send(message, message.text)


def process_and_send(message: types.Message, text: str):
    settings = get_settings(message.from_user.id)
    font_name = settings.get("font", list(FONTS.keys())[0])
    paper_kind = settings.get("paper", "daftar")

    font_file = FONTS[font_name]["file"]
    font_path = os.path.join(FONTS_DIR, font_file)

    if not os.path.exists(font_path):
        ensure_fonts()
    if not os.path.exists(font_path):
        bot.reply_to(
            message,
            f"❗ '{font_file}' shriftini internetdan yuklab bo'lmadi.\n"
            f"Iltimos, birozdan so'ng qayta urinib ko'ring yoki boshqa shrift tanlang.",
        )
        return

    bot.send_message(message.chat.id, "✍️ Navbatga qo'yildi, tez orada tayyor bo'ladi...")
    enqueue_job(message.chat.id, do_handwriting, message.chat.id, text, font_path, font_name, paper_kind)


def do_handwriting(chat_id: int, text: str, font_path: str, font_name: str, paper_kind: str):
    try:
        pages = render_pages(text=text, font_path=font_path, paper_kind=paper_kind)
        parts = export_parts(pages)
        total = len(parts)
        for i, (buf, ext, label) in enumerate(parts, start=1):
            buf.name = f"handwriting_{i}.{ext}" if total > 1 else f"handwriting.{ext}"
            caption = f"Shrift: {font_name}{label}"
            buf.seek(0)
            if ext == "jpg":
                send_with_retry(bot.send_photo, chat_id, buf, caption=caption)
            else:
                send_with_retry(bot.send_document, chat_id, buf, caption=caption)
    except Exception as e:
        logger.exception("Xatolik yuz berdi")
        bot.send_message(chat_id, f"❌ Xatolik: {e}")

# ============================================================
# ISHGA TUSHIRISH
# ============================================================
if __name__ == "__main__":
    os.makedirs(FONTS_DIR, exist_ok=True)
    os.makedirs(TEMPLATES_DIR, exist_ok=True)

    init_db()
    ensure_fonts()

    try:
        BOT_ID = bot.get_me().id
    except Exception as e:
        logger.warning(f"Bot ID olinmadi: {e}")

    threading.Thread(target=job_worker, daemon=True).start()
    start_keep_alive_server()
    start_self_ping()

    logger.info("Bot ishga tushdi...")
    bot.infinity_polling(skip_pending=True)
